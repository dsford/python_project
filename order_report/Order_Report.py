#-*- coding: utf-8 -*-

# Order Report 1612XX

import os
import re
import sys
import time
import xlrd
import codecs
import shutil
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

import dsford.backup
import dsford.ini
import dsford.search
import dsford.show
import dsford.time


# 프로그램 이름
program_name = "Order_Report"

# 프로그램 시작 시간
program_start_time = dsford.time.sub_folder_name()

# ini파일 로드
ini_dict = dsford.ini.load_ini(".\\Order_Report.ini")



# 리포트 관련 리스트
report_file_list = []
report_work_result = []
report_work_note = []


# 리포트 폴더 생성

if len(sys.argv) == 1:
    if ini_dict["ShowReport"].upper() == "True".upper():
        
        print("리포트 파일 생성 여부 : True")
        
        report_folder = ini_dict["Report"]
        if not os.path.isdir(report_folder): 
            os.mkdir(report_folder)

        report_file = open(report_folder + program_start_time + ".txt", "w")
        report_file.write("[작 업 결 과]\n\n")
        report_file.write("작업일시 : " + program_start_time + "\n\n")
        print(program_start_time + ".txt" + "파일을 생성했습니다.\n")


#카운트 리스트 생성
def make_count_list():
    count_list = []
    for i in range(int(ini_dict["MaxItemsCount"]) + 1):
        count_list.append("")
    return count_list



#번호에서 숫자만 추출
def conv_num(number):
 
    only_num = ini_dict["ExtractNumber"]
  
    num = re.findall(only_num, number)
  
    if len(str(num[0])) == 1:
      if ord(str(num[0])) >= ord("①"):
        num[0] = ord(str(num[0])) - ord("①") + 1
      
    return int(num[0])


#번호, 숫자만 남기고 나머지 지우기
def get_data(ori_text):

    text = re.sub(ini_dict["ItemSplit"], " ", ori_text).strip()
  
    number_pat = ini_dict["Number"]
    count_pat = ini_dict["Count"]

    get_list = []
    temp_list = []
    search_level = 0
  
    for word in text.split(" "):
      new_word = " " + word + " "
    
      temp_num = re.findall(number_pat, new_word)
      temp_cou = re.findall(count_pat, new_word)
         
      if search_level == 0:
        if len(temp_num) > 0 and len(temp_cou) > 0:
          temp_list.append(conv_num(temp_num[0]))
          temp_list.append(int(temp_cou[len(temp_cou) - 1]))
          search_level = 2       
        elif len(temp_num) > 0:        
          temp_list.append(conv_num(temp_num[0]))
          search_level = 1
        

      elif search_level > 0:      
        if len(temp_num) > 0 and len(temp_cou) > 0:
          if search_level == 2:
            get_list.append(temp_list)
            temp_list = []
            temp_list.append(conv_num(temp_num[0]))
            temp_list.append(int(temp_cou[len(temp_cou) - 1]))          
          elif search_level == 1:
            temp_list.append(int(temp_cou[len(temp_cou) - 1]))
            search_level = 2
          
        elif len(temp_num) > 0:
          if search_level == 2:
            get_list.append(temp_list)
            temp_list = []
            temp_list.append(conv_num(temp_num[0]))
            search_level = 1
          
        elif len(temp_cou) > 0:     
          if search_level == 2:  
            temp_list[1] = int(temp_cou[len(temp_cou) - 1])
          elif search_level == 1:
            temp_list.append(int(temp_cou[len(temp_cou) - 1]))
            search_level = 2
          
  
    get_list.append(temp_list)           
            

    
    count_list = make_count_list()
    for i in range(len(get_list)):
        if count_list[get_list[i][0]] == "":
            count_list[get_list[i][0]] = 0
        count_list[get_list[i][0]] = count_list[get_list[i][0]] + get_list[i][1]
        
    
    
    temp_sum = 0
    
    for i in range(1, len(count_list)):
        if not count_list[i] == "":
            temp_sum = temp_sum + count_list[i]

    count_list[0] = temp_sum
    
    
    return count_list


# 숫자를 열번호로 변경 (1 => A)
def num2col(number):


    ascii_string = [chr(i) for i in range(65, 91)]

    
    string = []

    under = 0
    if number > 26:
        under = 1

    while True:

        number, r = divmod(number, 26)

    
        if r == 0:
            r = 26
            number -= 1

        string.append(ascii_string[r - 1])
       

        if number < 26:
            if not number == 0:
                string.append(ascii_string[number - 1])
            
            new_string = []
            for i in range(len(string)-1,-1,-1):
                new_string.append(string[i])
            
            return "".join(new_string)             
 
            break
        


#xlsx파일에 직접 테이블 기록
def write_xlsx(wb2, ws2, file, ncol, trow, total_count_list, new_max_order_count, first_col, is_filter):

    HeadColColor = str(ini_dict["HeadColColor"])
    TotalRowColor = str(ini_dict["TotalRowColor"])
    
    #합계위치옵션 생략으로 인한 수정(161205) => 다시 부활(161206)
    
    for i in range(new_max_order_count + 1):
        if not i == 0:
            ws2.cell(column=ncol + 1 + i, row=trow + 1, value="옵션" + str(i))
          
        else:
            ws2.cell(column=ncol + 1 + i, row=trow + 1, value="합계")

        ws2.cell(column=ncol + 1 + i, row=trow + 1).alignment=Alignment(horizontal="center",vertical="center")
        ws2.cell(column=ncol + 1 + i, row=trow + 1).font = Font(bold=True)
        ws2.cell(column=ncol + 1 + i, row=trow + 1).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws2.cell(column=ncol + 1 + i, row=trow + 1).fill = PatternFill(fill_type="solid", start_color=HeadColColor)

            
    for i in range(len(total_count_list)):
        for j in range(new_max_order_count + 1):
            if not total_count_list[i][j] == "":
                ws2.cell(column=ncol + 1 + j, row=trow + 2 + i, value=int(total_count_list[i][j]))

            ws2.cell(column=ncol + 1 + j, row=trow + 2 + i).alignment=Alignment(horizontal="center",vertical="center")
            ws2.cell(column=ncol + 1 + j, row=trow + 2 + i).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

            if j == 0:
                ws2.cell(column=ncol + 1 + j, row=trow + 2 + i).font = Font(bold=True)
                ws2.cell(column=ncol + 1 + j, row=trow + 2 + i).fill = PatternFill(fill_type="solid", start_color=TotalRowColor)


            if i == len(total_count_list) - 1:
                ws2.cell(column=ncol + 1 + j, row=trow + 2 + i).font = Font(bold=True)
                ws2.cell(column=ncol + 1 + j, row=trow + 2 + i).fill = PatternFill(fill_type="solid", start_color=TotalRowColor)


    if is_filter == 1:
        range_address = num2col(first_col + 1) + str(trow + 1) + ":" + num2col(ncol + 1 + new_max_order_count - 1) + str(trow + 2 + len(total_count_list) - 1)
        sort_range_address = num2col(ncol + 1 + 1) + str(trow + 1 + 1) + ":" + num2col(ncol + 1 + 1 - 1) + str(trow + 2 + len(total_count_list) - 1)
        ws2.auto_filter.ref = range_address
            
            
    wb2.save(filename = file)

    print("\n" + file.split("\\")[len(file.split("\\")) - 1] + "에 대한 작업을 완료했습니다.\n")
    print("처리할 다른 파일이 남아있다면 2초 후 다시 시작합니다.\n")
    time.sleep(2)
    if ini_dict["ShowReport"].upper() == "True".upper():
        report_work_result.append("성공")
        report_work_note.append("")



    #항목합계 생략
    """
    if ManageDataOrderBy.upper() == "TotalToCase".upper():
        for i in range(new_max_order_count + 1):
            if not i == 0:
                ws2.cell(column=ncol + 1 + i, row=trow + 1, value="옵션" + str(i))
            else:
                ws2.cell(column=ncol + 1 + i, row=trow + 1, value="합계")
            ws2.cell(column=ncol + 1 + i, row=trow + 1).alignment=Alignment(horizontal="center",vertical="center")
            ws2.cell(column=ncol + 1 + i, row=trow + 1).font = Font(bold=True)
            ws2.cell(column=ncol + 1 + i, row=trow + 1).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            ws2.cell(column=ncol + 1 + i, row=trow + 1).fill = PatternFill(fill_type="solid", start_color=HeadColColor)

        for i in range(len(total_count_list)):
            for j in range(new_max_order_count + 1):
                if not total_count_list[i][j] == "":
                    ws2.cell(column=ncol + 1 + j, row=trow + 2 + i, value=int(total_count_list[i][j]))

                ws2.cell(column=ncol + 1 + j, row=trow + 2 + i).alignment=Alignment(horizontal="center",vertical="center")
                ws2.cell(column=ncol + 1 + j, row=trow + 2 + i).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

                if j == 0:
                    ws2.cell(column=ncol + 1 + j, row=trow + 2 + i).font = Font(bold=True)
                    ws2.cell(column=ncol + 1 + j, row=trow + 2 + i).fill = PatternFill(fill_type="solid", start_color=TotalRowColor)


        if is_filter == 1:
            range_address = num2col(first_col + 1) + str(trow + 1) + ":" + num2col(ncol + 1 + new_max_order_count) + str(trow + 2 + len(total_count_list) - 1)
            sort_range_address = num2col(ncol + 1 + 1) + str(trow + 1 + 1) + ":" + num2col(ncol + 1 + 1) + str(trow + 2 + len(total_count_list) - 1)
            ws2.auto_filter.ref = range_address
            
            
        wb2.save(filename = file)

        print("\n" + file.split("\\")[len(file.split("\\")) - 1] + "에 대한 작업을 완료했습니다.\n")
        print("처리할 다른 파일이 남아있다면 2초 후 다시 시작합니다.\n")
        time.sleep(2)
        if ini_dict["ShowReport"].upper() == "True".upper():
            report_work_result.append("성공")
            report_work_note.append("")

    elif ManageDataOrderBy.upper() == "CaseToTotal".upper():
        for i in range(new_max_order_count + 1):
            if not i == new_max_order_count:
                ws2.cell(column=ncol + 1 + i, row=trow + 1, value="옵션" + str(i + 1))
            else:
                ws2.cell(column=ncol + 1 + i, row=trow + 1, value="합계")
            ws2.cell(column=ncol + 1 + i, row=trow + 1).alignment=Alignment(horizontal="center",vertical="center")
            ws2.cell(column=ncol + 1 + i, row=trow + 1).font = Font(bold=True)
            ws2.cell(column=ncol + 1 + i, row=trow + 1).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            ws2.cell(column=ncol + 1 + i, row=trow + 1).fill = PatternFill(fill_type="solid", start_color=HeadColColor)


        for i in range(len(total_count_list)):
            for j in range(1, new_max_order_count + 1):
                if not total_count_list[i][j] == "":
                    ws2.cell(column=ncol + j, row=trow + 2 + i, value=int(total_count_list[i][j]))
   
                ws2.cell(column=ncol + j, row=trow + 2 + i).alignment=Alignment(horizontal="center",vertical="center")
                ws2.cell(column=ncol + j, row=trow + 2 + i).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

            if not total_count_list[i][0] == "":
                ws2.cell(column=ncol + new_max_order_count + 1, row=trow + 2 + i, value=int(total_count_list[i][0]))
            ws2.cell(column=ncol + new_max_order_count + 1, row=trow + 2 + i).alignment=Alignment(horizontal="center",vertical="center")
            ws2.cell(column=ncol + new_max_order_count + 1, row=trow + 2 + i).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            ws2.cell(column=ncol + new_max_order_count + 1, row=trow + 2 + i).font = Font(bold=True)
            ws2.cell(column=ncol + new_max_order_count + 1, row=trow + 2 + i).fill = PatternFill(fill_type="solid", start_color=TotalRowColor)


        if is_filter == 1:
            range_address = num2col(first_col + 1) + str(trow + 1) + ":" + num2col(ncol + 1 + new_max_order_count) + str(trow + 2 + len(total_count_list) - 1)
            sort_range_address = num2col(ncol + 1 + 1) + str(trow + 1 + 1) + ":" + num2col(ncol + 1 + 1) + str(trow + 2 + len(total_count_list) - 1)
            ws2.auto_filter.ref = range_address
            
                    
        wb2.save(filename = file)

    
        print("\n" + file.split("\\")[len(file.split("\\")) - 1] + "에 대한 작업을 완료했습니다.\n")
        print("처리할 다른 파일이 남아있다면 2초 후 다시 시작합니다.\n")
        time.sleep(2)
        if ini_dict["ShowReport"].upper() == "True".upper():
            report_work_result.append("성공")
            report_work_note.append("")
   

    
    else:
        os.system("cls")
        print("\nManageDataOrderBy 옵션이 올바르지 않습니다.")
        print("프로그램을 종료합니다. ini파일을 확인해 주세요.")
        time.sleep(2)
        if ini_dict["ShowReport"].upper() == "True".upper():
            report_file.write("ManageDataOrderBy 설정이 올바르지 않습니다.\n")
            report_file.close()
            os.startfile(report_folder + program_start_time + ".txt")
                
        sys.exit()

    """



# 파일에서 데이터 추출
def data_summary(file):
    

    print(file.split("\\")[len(file.split("\\")) - 1] + " 작업 시작")
    
    #파일 읽기
    wb = xlrd.open_workbook(file)
    ws = wb.sheet_by_index(0)

    #머리글 행 검색
    ncol = ws.ncols
    nrow = ws.nrows
    first_col = 0

    tcol = 0
    trow = 0

    temp_col = []
    temp_row = []

    title_name_db = []

    field_num = 1
    while True:    
        try:
            field = ini_dict["Field" + str(field_num)]
            if not field == "":
                title_name_db.append(field)
            field_num += 1
        except:
            break


    loop_end = 0
    sub_loop_end = 0
    temp_i = 0
    while loop_end == 0:
        temp = ws.row_values(temp_i)
        
        for i in range(len(temp)):
            for j in title_name_db:
                if temp[i] == j:
                    temp_col.append(i)
                    temp_row.append(temp_i)

                    for k in range(len(temp)):
                        if not temp[k] == "":
                            first_col = k
                            break

                    
                    sub_loop_end = 1
                    
        if sub_loop_end == 1:
            loop_end = 1

            if len(temp_col) == 1:
                tcol = temp_col[0]
                trow = temp_row[0]
                

            else:                

                ManageHeaderColMultipleResult = ini_dict["ManageHeaderColMultipleResult"]
                
                if  ManageHeaderColMultipleResult.upper() == "Ahead".upper():
                    tcol = temp_col[0]
                    trow = temp_row[0]
                    
                elif ManageHeaderColMultipleResult.upper() == "Last".upper():
                    tcol = temp_col[len(temp_col) - 1]
                    trow = temp_row[len(temp_row) - 1]

                elif ManageHeaderColMultipleResult.upper() == "Select".upper():

                    manage_loop = 0
                    while manage_loop == 0:
                        os.system('cls')
                        print("\n대상파일 : " + file.split("\\")[len(file.split("\\")) - 1])
                        print(str(len(temp_col)) + "개의 머리글 행이 검색되었습니다.\n\n\n")

                    
                        for i in range(len(temp_col)):
                            temp_db = ws.col_values(temp_col[i])
                            
                            #print("[" + str(i + 1) + "]    " + temp_db[int(len(temp_db) / 2) + 1][:50] + "\n\n")
                            print("[" + str(i + 1) + "]    " + str(temp_db[temp_row[i] + 1])[:50] + "\n\n")
                            #항목 바로 아래 데이터를 가져오도록 수정(161205)


                        print("[X]    프로그램 종료(파일이 여러개면 다음 파일로 넘어갑니다.)\n\n")
                            
                        manage_select = input("\n데이터가 있는 항목의 숫자를 입력하고 Enter를 눌러 주세요 : ")
                        
                        if manage_select.upper() == "X":
                            report_work_result.append("실패")
                            report_work_note.append("사용자가 임의로 작업 종료")
                            return None
                                               
                              
                        for i in range(len(temp_col)):
                            if manage_select == str(i + 1): 
                                tcol = temp_col[i]
                                trow = temp_row[i]
                                manage_loop = 1
                                os.system('cls')
                                break
                        
                        if manage_loop == 0:
                            os.system('cls')
                            print("\n잘못 입력하셨습니다. 1초 후 선택 화면으로 돌아갑니다.")
                            time.sleep(1)

            
        if loop_end == 1:
            break
        temp_i = temp_i + 1
        if temp_i > nrow - 1:
            loop_end = 1
            
            print("\n대상파일 : " + file.split("\\")[len(file.split("\\")) - 1] + "\n")
            print("주문관련 머리글 행이 없습니다. 파일을 확인해 주세요.\n")
            print("처리할 다른 파일이 남아있다면 2초 후 다시 시작합니다.\n")

            report_work_result.append("실패")
            report_work_note.append("주문관련 머리글 행 없음")
            
            time.sleep(2)
            return None


    #번호, 숫자 취합 작업
    data = ws.col_values(tcol)

    total_count_list = []
    for i in range(trow + 1, len(data)):
        
        if data[i] == "":
            total_count_list.append(make_count_list())
        else:
            total_count_list.append(get_data(data[i]))

    #옵션합계(161205)
    option_sum = []
    for i in range(len(total_count_list[0])):
        option_temp_sum = 0
        
        for j in range(len(total_count_list)):
            try:
                option_temp_sum = option_temp_sum + int(total_count_list[j][i])
            except:
                pass          
        if option_temp_sum == 0:
            option_temp_sum = ""
        option_sum.append(option_temp_sum)
        
    total_count_list.append(option_sum)
    
   

    temp_max = 1
    for i in range(trow + 1, len(total_count_list)):
        for j in range(len(total_count_list[i])):
            if not total_count_list[i][j] == "":
                temp_max = max(temp_max, j)
    
    new_max_order_count = temp_max

    
                

    #파일쓰기 작업

    #ManageDataOrderBy = ini_dict["ManageDataOrderBy"]

    if file[-4:].upper() == "xlsx".upper():

        wb2 = openpyxl.load_workbook(file)
        ws2 = wb2.active

        is_filter = 0
        if ini_dict["ManageHeaderColFilters"].upper() == "True".upper():
            is_filter = 1
    
        write_xlsx(wb2, ws2, file, ncol, trow, total_count_list, new_max_order_count, first_col, is_filter)
        

    elif file[-3:].upper() == "xls".upper():

        print("\n대상파일 : " + file.split("\\")[len(file.split("\\")) - 1])
        print("해당 파일은 .xls 이므로 정해둔 설정에 따라 파일을 따로 생성합니다.")
        
        ManageXlsDataWrite = ini_dict["ManageXlsDataWrite"]

        if ManageXlsDataWrite.upper() == "Excel".upper():
           
            
            wb2 = openpyxl.Workbook()
            ws2 = wb2.active

            new_file = file[:-4] + "_Report.xlsx"
        
            write_xlsx(wb2, ws2, new_file, 0, trow, total_count_list, new_max_order_count, 0, 0)

    
        elif ManageXlsDataWrite.upper() == "Txt".upper():

            txt_file = open(file[:-3] + "txt", "w")


            for i in range(new_max_order_count + 1):
                if i == 0:
                    txt_file.write("합계" + "\t")
                    
                else:
                    txt_file.write("옵션" + str(i) + "\t")
                    
            txt_file.write("\r\n")

            for i in range(len(total_count_list)):
                
                #합계랑 구분하기 위한 선 그리기 
                if i == len(total_count_list) - 1:
                    txt_file.write("=" * (new_max_order_count + 1) * 7 + "\r\n")
                        
                for j in range(new_max_order_count + 1):                                
                    txt_file.write(str(total_count_list[i][j]) + "\t")
                txt_file.write("\r\n")

            txt_file.close()
            print(file.split("\\")[len(file.split("\\")) - 1][:-3] + "txt" + "파일을 생성했습니다.\n")
            report_work_result.append("성공")
            report_work_note.append(file.split("\\")[len(file.split("\\")) - 1][:-3] + "txt" + "로 생성")


            # ManageDataOrderBy 옵션 미사용으로 인한 생략(1612XX)
            """

            if ManageDataOrderBy.upper() == "TotalToCase".upper():
                for i in range(new_max_order_count + 1):
                    if i == 0:
                        pass
                        #txt_file.write("합계" + "\t")
                    else:
                        txt_file.write("옵션" + str(i) + "\t")
                txt_file.write("\r\n")

                for i in range(len(total_count_list)):
                    for j in range(new_max_order_count + 1):
                        #합계랑 구분하기 위한 선 그리기
                        if i == len(total_count_list) - 1:
                            txt_file.write("=" * new_max_order_count * 2 + "\r\n")
                        if not j == 0:
                            txt_file.write(str(total_count_list[i][j]) + "\t")
                    txt_file.write("\r\n")

                txt_file.close()
                print(file.split("\\")[len(file.split("\\")) - 1][:-3] + "txt" + "파일을 생성했습니다.\n")
                report_work_result.append("성공")
                report_work_note.append(file.split("\\")[len(file.split("\\")) - 1][:-3] + "txt" + "로 생성")

            elif ManageDataOrderBy.upper() == "CaseToTotal".upper():
                for i in range(new_max_order_count + 1):
                    if not i == new_max_order_count:
                        txt_file.write("옵션" + str(i + 1) + "\t")
                    else:
                        txt_file.write("합계" + "\t")
                txt_file.write("\r\n")

                for i in range(len(total_count_list)):
                    for j in range(1, new_max_order_count + 1):
                        txt_file.write(str(total_count_list[i][j]) + "\t")
                    txt_file.write(str(total_count_list[i][0]) + "\t")
                    txt_file.write("\r\n")

                txt_file.close()
                print("\n" + file.split("\\")[len(file.split("\\")) - 1][:-3] + "txt" + "에 대한 작업을 완료했습니다.\n")
                report_work_result.append("성공")
                report_work_note.append(file.split("\\")[len(file.split("\\")) - 1][:-3] + "txt" + "로 생성")

            else:
                os.system("cls")
                print("\nManageDataOrderBy 옵션이 올바르지 않습니다.")
                print("프로그램을 종료합니다. ini파일을 확인해 주세요.")
                time.sleep(2)
                if ini_dict["ShowReport"].upper() == "True".upper():
                    report_file.write("ManageDataOrderBy 설정이 올바르지 않습니다.\n")
                    report_file.close()
                    os.startfile(report_folder + program_start_time + ".txt")
                
                sys.exit()
            """
        else:
                os.system("cls")
                print("\nManageXlsDataWrite 설정이 올바르지 않습니다.\n")
                print("프로그램을 종료합니다. ini파일을 확인해 주세요.")           
                time.sleep(2)
                if ini_dict["ShowReport"].upper() == "True".upper():
                    report_file.write("ManageXlsDataWrite 설정이 올바르지 않습니다.\n")
                    report_file.close()
                    os.startfile(report_folder + program_start_time + ".txt")
                
                sys.exit()
    
    else:
        os.system("cls")
        print("\nManageXlsDataWrite 옵션이 올바르지 않습니다.\n")
        print("프로그램을 종료합니다. ini파일을 확인해 주세요.")
        time.sleep(2)
        if ini_dict["ShowReport"] == "True":
            report_file.write("ManageXlsDataWrite 설정이 올바르지 않습니다.\n")
            report_file.close()
            os.startfile(report_folder + program_start_time + ".txt")
                
        sys.exit()                      
    
    






# (반)자동 모드
def auto_mode(file_list, full_file_list):
    if ini_dict["ManageHeaderColMultipleResult"].upper() == "Select".upper():
        
         print("작동 모드 : 반자동 (사용자가 입력해야 할 작업이 있을 수 있습니다.)\n ")

    else:

        print("작동 모드 : 자동 (미리 정해둔 설정대로 작업합니다.)\n\n")

    mode_delay = int(ini_dict["ModeDelay"])
    print("\n" + str(mode_delay) + "초 후 작업을 시작합니다.\n")
    time.sleep(mode_delay)


    for file in full_file_list:
        data_summary(file)



def main():

    if len(sys.argv) > 1:
        if sys.argv[1] == '-rp':
            print("xptmxm")
            sys.exit()
        else:
            print("너무 많은 인수를 입력하셨습니다.\n프로그램을 종료합니다.")
            sys.exit()

    # 경구문구 출력
    if ini_dict["ShowWarning"].upper() == "True".upper():
        dsford.show.warning(program_name, int(ini_dict["ShowWarningDelay"]))



    # .xlsx .xls 확장자를 가진 파일 목록 생성
    file_ext = []
    ext_num = 1
    while True:    
        try:
            ext = ini_dict["Ext" + str(ext_num)]
            if not ext == "":
                file_ext.append(ext)
            ext_num += 1
        except:
            break

    file_list = []
    full_file_list = []

    for i in file_ext:

        file_list += dsford.search.ext(ini_dict["File"], i, 0)
        full_file_list += dsford.search.ext(ini_dict["File"], i, 1)

    if len(file_list) == 0:
        print(base_dir + "에 대상 파일이 존재하지 않습니다.")
        print("프로그램을 종료합니다.")       

        if load_ini()["ShowReport"].upper() == "True".upper():
            report_file.write(base_dir + " 에 대상 파일이 존재하지 않습니다.\n")
            report_file.close()

            os.startfile(report_folder + program_start_time + ".txt")
        
        time.sleep(2)
        
        sys.exit()

    print("\n대상 파일 : " + str(len(file_list)) + "개")

    for i in file_list:
        report_file_list.append(i)

    
    # 파일 백업
    if ini_dict["DoFileBackup"].upper() == "True".upper():
        
        print("파일 백업 여부 : True")

        for i in range(len(file_list)):

            dsford.backup.backup(ini_dict["Backup"], program_start_time, file_list[i], full_file_list[i])

        print(ini_dict["Backup"] + '\\' + program_start_time + '\\' + " 에 파일을 백업했습니다.\n\n")
        


    # 모드에 따라 실행 // 일단은 반자동 모드로

    excute_mode = ini_dict["Mode"]

    if excute_mode.upper() == "Auto".upper():

        auto_mode(file_list, full_file_list)

        print("\n\n모든 작업이 완료되었습니다.")
        if ini_dict["ShowReport"].upper() == "True".upper():
            for i in range(len(report_file_list)):
                report_file.write("  " + report_file_list[i] + " : " + report_work_result[i])
                if not report_work_note[i] == "":
                    report_file.write(" // " + report_work_note[i])
                report_file.write("\n")
                
            report_file.close()

            os.startfile(report_folder + program_start_time + ".txt")

    else:

        print("\nMode 설정이 올바르지 않습니다.\n")
        print("프로그램을 종료합니다.")       

        if ini_dict["ShowReport"].upper() == "True".upper():
            report_file.write(base_dir + " 에 대상 파일이 존재하지 않습니다.\n")
            report_file.close()

            os.startfile(report_folder + program_start_time + ".txt")
        
        time.sleep(2)
        
        sys.exit()
        

if __name__ == '__main__':
    sys.exit(main())
