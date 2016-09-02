#-*- coding: utf-8 -*-
import os
import re
import sys
import time
import xlrd
import codecs
import shutil
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


program_version = "v160804"

time_delay = 1

next_time = []
note = []





# 경고문

def warning():

    os.system("cls")

    print("\nSearch Time " + program_version)
    print("         by D.S.Ford\n\n")

    print("\n[주 의 사 항]\n")
    print("이 프로그램은 제한된 정보만을 기준으로 만들었기 때문에")
    print("경우에 따라서 오류가 발생하거나 제대로된 결과물을 낼 수 없을 수도 있습니다.")
    print("따라서 결과물을 100% 신뢰해서는 안 되며, 반드시 직접 확인하셔야 합니다.\n\n")
    print("[확 인 사 항]\n")
    print("1) xlsx 파일만을 사용해야 합니다.")
    print("2) 데이터는 시간-시가-고가-저가-종가 순이어야 합니다.")
    
    print("2) 종가가 있는 열 기준, \"우측 두 개의 열\"을 사용하므로")
    print("   이 곳에 다른 데이터가 없도록 해야 합니다.")
    print("3) 한 파일 내에 관련 데이터가 여러 시트에 있더라도")
    print("   가장 먼저 나오는 시트 하나만을 대상으로 작업합니다.\n\n\n")
    
    print("5초 후 프로그램을 시작합니다.")

    time.sleep(5)

    os.system("cls")


    
#대상 파일 검색            
def search():

    file_list = [] 
         
    file_name = os.listdir(".\\")
        
    for file in file_name:
        ext = os.path.splitext(file)[-1]
        if ext == ".xlsx": 
            file_list.append(file)
            
    return file_list




def sub_loop(temp_loop, loop, time_row, time_table, cur_table, high_table, low_table, end_table, base_data, temp_sum):

    for j in range(loop+1, len(cur_table)):

        if abs(float(base_data) - float(cur_table[j])) >= 1.25:
            next_time.append(time_table[j])
            if temp_sum == "01":
                note.append("저가 // " + str(time_row + j + 1) + "행 시가")
            elif temp_sum == "10":
                note.append("고가 // " + str(time_row + j + 1) + "행 시가")
            for k in range(loop+1, j+1):
                next_time.append("")
                note.append("")
            loop = j + 1
            return loop
            break
                
        elif abs(float(base_data) - float(high_table[j])) >= 1.25:
            next_time.append(time_table[j])
            if temp_sum == "01":
                note.append("저가 // " + str(time_row + j + 1) + "행 고가")
            elif temp_sum == "10":
                note.append("고가 // " + str(time_row + j + 1) + "행 고가")               
            for k in range(loop+1, j+1):
                next_time.append("")
                note.append("")
            loop = j + 1
            return loop
            break
                
        elif abs(float(base_data) - float(low_table[j])) >= 1.25:
            next_time.append(time_table[j])
            if temp_sum == "01":
                note.append("저가 // " + str(time_row + j + 1) + "행 저가")
            elif temp_sum == "10":
                note.append("고가 // " + str(time_row + j + 1) + "행 저가")             
            for k in range(loop+1, j+1):
                next_time.append("")
                note.append("")
            loop = j + 1
            return loop
            break
                
        elif abs(float(base_data) - float(end_table[j])) >= 1.25:
            next_time.append(time_table[j])
            if temp_sum == "01":
                note.append("저가 // " + str(time_row + j + 1) + "행 종가")
            elif temp_sum == "10":
                note.append("고가 // " + str(time_row + j + 1) + "행 종가")                 
            for k in range(loop+1, j+1):
                next_time.append("")
                note.append("")
            loop = j + 1
            return loop
            break
    
    if temp_loop == loop:
            next_time.append("")
            if temp_sum == "01":
                note.append("저가(없음)")
            elif temp_sum == "10":
                note.append("고가(없음)")
            loop += 1
            return loop



def sub_loop_two(temp_loop, loop, time_row, time_table, cur_table, high_table, low_table, end_table):

    for j in range(loop+1, len(cur_table)):

        if abs(float(high_table[loop]) - float(cur_table[j])) >= 1.25:
            next_time.append(time_table[j])
            note.append("고가(동) // " + str(time_row + j + 1) + "행 시가")
            for k in range(loop+1, j+1):
                next_time.append("")
                note.append("")
            loop = j + 1
            return loop
            break


        elif abs(float(low_table[loop]) - float(cur_table[j])) >= 1.25:
            next_time.append(time_table[j])
            note.append("저가(동) // " + str(time_row + j + 1) + "행 시가")
            for k in range(loop+1, j+1):
                next_time.append("")
                note.append("")
            loop = j + 1
            return loop
            break


        
                
        elif abs(float(high_table[loop]) - float(high_table[j])) >= 1.25:
            next_time.append(time_table[j])
            note.append("고가(동) // " + str(time_row + j + 1) + "행 고가")               
            for k in range(loop+1, j+1):
                next_time.append("")
                note.append("")
            loop = j + 1
            return loop
            break

        elif abs(float(low_table[loop]) - float(high_table[j])) >= 1.25:
            next_time.append(time_table[j])
            note.append("저가(동) // " + str(time_row + j + 1) + "행 고가")               
            for k in range(loop+1, j+1):
                next_time.append("")
                note.append("")
            loop = j + 1
            return loop
            break


           
        elif abs(float(high_table[loop]) - float(low_table[j])) >= 1.25:
            next_time.append(time_table[j])
            note.append("고가(동) // " + str(time_row + j + 1) + "행 저가")             
            for k in range(loop+1, j+1):
                next_time.append("")
                note.append("")
            loop = j + 1
            return loop
            break

        elif abs(float(low_table[loop]) - float(low_table[j])) >= 1.25:
            next_time.append(time_table[j])
            note.append("저가(동) // " + str(time_row + j + 1) + "행 저가")             
            for k in range(loop+1, j+1):
                next_time.append("")
                note.append("")
            loop = j + 1
            return loop
            break

 
                
        elif abs(float(high_table[loop]) - float(end_table[j])) >= 1.25:
            next_time.append(time_table[j])
            note.append("고가(동) // " + str(time_row + j + 1) + "행 종가")                 
            for k in range(loop+1, j+1):
                next_time.append("")
                note.append("")
            loop = j + 1
            return loop
            break

        elif abs(float(low_table[loop]) - float(end_table[j])) >= 1.25:
            next_time.append(time_table[j])
            note.append("저가(동) // " + str(time_row + j + 1) + "행 종가")                 
            for k in range(loop+1, j+1):
                next_time.append("")
                note.append("")
            loop = j + 1
            return loop
            break

        
    
    if temp_loop == loop:
            next_time.append("")
            note.append("동시만족(없음)")
            loop += 1
            return loop



def data_analysis(wb, ws, nrows, ncols, time_row, time_col):


    time_table = ws.col_values(time_col)
    cur_table = ws.col_values(time_col+1)
    high_table = ws.col_values(time_col+2)
    low_table = ws.col_values(time_col+3)
    end_table = ws.col_values(time_col+4)

    

    loop = time_row + 1
    
    while loop < len(cur_table):
        
        temp_loop = loop

        temp = ["0", "0"]

        
        if not cur_table[loop] == "":
            if abs(float(cur_table[loop]) - float(high_table[loop])) >= 1.25:
                temp[0] = "1"
            if abs(float(cur_table[loop]) - float(low_table[loop])) >= 1.25:
                temp[1] = "1"
    
            temp_sum = "".join(temp)

            if temp_sum == "00":
                next_time.append("")
                note.append("")
                loop += 1


            elif temp_sum == "01":
            
                loop = sub_loop(temp_loop, loop, time_row, time_table, cur_table, high_table, low_table, end_table, low_table[loop], temp_sum)


            elif temp_sum == "10":
            
                loop = sub_loop(temp_loop, loop, time_row, time_table, cur_table, high_table, low_table, end_table, high_table[loop], temp_sum)


            elif temp_sum == "11":

                loop = sub_loop_two(temp_loop, loop, time_row, time_table, cur_table, high_table, low_table, end_table) 
        else:
            loop += 1

    return None



def writing_to_file(file, full_file, sheet_name, nrows, ncols, time_row, time_col):

    wb2 = openpyxl.load_workbook(full_file)
    ws2 = wb2.get_sheet_by_name(sheet_name)

    time_row += 1
    time_col += 1

    next_time_col = time_col + 5
    note_col = time_col + 6


    ws2.cell(column=next_time_col, row=time_row, value="만족 시간").alignment=Alignment(horizontal="center",vertical="center")
    ws2.cell(column=note_col, row=time_row, value="비 고").alignment=Alignment(horizontal="center",vertical="center")
    ws2.cell(column=next_time_col, row=time_row).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    ws2.cell(column=note_col, row=time_row).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    


    for i in range(len(next_time)):
        ws2.cell(column=next_time_col, row=time_row+1+i, value=next_time[i]).alignment=Alignment(horizontal="center",vertical="center")
        ws2.cell(column=note_col, row=time_row+1+i, value=note[i]).alignment=Alignment(horizontal="left",vertical="center")
        
        ws2.cell(column=next_time_col, row=time_row+1+i).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws2.cell(column=note_col, row=time_row+1+i).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))



    dims = {}
    for row in ws2.rows:
        for cell in row:
            if cell.value == "비 고":
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws2.column_dimensions[col].width = value + 25
 
    
        

    wb2.save(filename = full_file)






def main():

    warning()

    # xlsx 파일 검색
    file = search()

    # 2개 이상 나왔을 경우 처리
    if len(file) > 1:
        
        manage_loop = 0
        while manage_loop == 0:
            os.system("cls")

            print("\nxlsx 파일 " + str(len(file)) + "개가 발견되었습니다.\n\n")
            
            for i in range(len(file)):
                print("[" + str(i + 1) + "] " + file[i] + "\n")
            print("[X] 프로그램 종료\n")

            manage_select = input("\n데이터가 있는 항목의 숫자를 입력하고 Enter를 눌러 주세요 : ")

            if manage_select.upper() == "X":
                print("\n프로그램을 종료합니다.")
                time.sleep(time_delay)
                sys.exit()
                
            for i in range(len(file)):
                if manage_select == str(i + 1):
                    file = file[i]
                    print("\n")
                    manage_loop = 1
                    break
                
            if manage_loop == 0:
                print("\n잘못 입력하셨습니다. 1초 후 선택 화면으로 돌아갑니다.")
                time.sleep(time_delay)
    elif len(file) == 1:
        file = file[0]

    else:
        print("\nxlsx 파일이 없습니다. 프로그램을 종료합니다.")
        time.sleep(time_delay)
        sys.exit()

    # 전체 경로 (openpyxl 전용)
    full_file = os.path.abspath(os.curdir) + "\\" + file

   


    # 파일 백업
    backup_folder = ".\\Backup"

    if not os.path.isdir(backup_folder): 
            os.mkdir(backup_folder)
    
    program_start_time = "%04d-%02d-%02d %02d%02d%02d" % (time.localtime().tm_year, time.localtime().tm_mon, time.localtime().tm_mday, time.localtime().tm_hour, time.localtime().tm_min, time.localtime().tm_sec)

    backup_sub_folder = program_start_time
    if not os.path.isdir(backup_folder + '\\' + backup_sub_folder): 
            os.mkdir(backup_folder + '\\' + backup_sub_folder)
            
    shutil.copy(full_file, backup_folder + '\\' + backup_sub_folder + '\\' + file)

    print("\n" + backup_folder + '\\' + backup_sub_folder + '\\' + " 에 [" + file + "] 파일을 백업했습니다.\n\n")
    time.sleep(time_delay)
    


    # 파일 기본 정보 획득
    print("[대상 파일 : " + file + "]\n")

    print("파일 분석을 시작합니다.\n")

    wb = xlrd.open_workbook(file)
    sc = wb.sheet_names()

    nrows = 0
    ncols = 0

    time_row = 0
    time_col = 0
    temp_end = 0

    sheet_name = ""
    
    for i in range(len(sc)):

        print("[" + sc[i] + "] 시트를 검사중입니다.")
        
        ws = wb.sheet_by_name(sc[i])

        nrows = ws.nrows
        ncols = ws.ncols

        for j in range(nrows):
            temp = ws.row_values(j)

            for k in range(len(temp)):
                if temp[k] == "시간":
                    time_row = j
                    time_col = k
                    temp_end = 1
                    break
            if not temp_end == 0:
                break

        if not temp_end == 0:
            print("[" + sc[i] + "] 시트에서 데이터를 찾았습니다.\n")
            sheet_name = sc[i]
            break

        print("[" + sc[i] + "] 시트에 데이터가 없습니다.\n")

        if sheet_name == "":
            print("이 파일에는 데이터가 없습니다.\n\n프로그램을 종료합니다.")
            time.sleep(time_delay + 2)
            sys.exit()


    # 데이터 분석
    print("데이터를 분석중입니다...")
        
    data_analysis(wb, ws, nrows, ncols, time_row, time_col)        
    
    print("데이터 분석을 완료했습니다.\n")



    # 파일 쓰기
    print("분석 결과를 파일에 입력중입니다...")

    writing_to_file(file, full_file, sheet_name, nrows, ncols, time_row, time_col)

    print("모든 결과를 입력 완료했습니다.\n\n")
    print("프로그램을 종료합니다.")
    time.sleep(time_delay + 2)
    sys.exit()




if __name__ == '__main__':
    sys.exit(main())
