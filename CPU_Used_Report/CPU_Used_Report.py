#-*- coding: utf-8 -*-

# CPU_Used_Report v161101

import os
import re
import sys
import time
import codecs
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


# 프로그램 이름
program_name = "CPU_Used_Report"

# 텍스트 파일 위치할 경로
base_dir = ".\\"
ext = "txt"



def load_ini(ini_file_dir):

    ini_file = codecs.open(ini_file_dir, "r", "utf-16")

    ini_dict = {}

    for line in ini_file:
        temp = line.strip()
        if not temp == "" and not temp[0] == "[" and not temp[0] == "#":
            temp = line.strip().replace("==", "=▣")
            temp = temp.split("=")
            temp[1] = temp[1].replace("▣", "=")

            try:
                ini_dict[temp[0]] = int(temp[1])
            except:
                ini_dict[temp[0]] = temp[1]

    ini_file.close()
          
    return ini_dict


# 설정 파일 불러오기
ini_file = load_ini(base_dir + program_name + ".ini")

# 업무시간 구분
start_working_time = time.strptime(ini_file["업무시작시간"], "%H:%M:%S")
end_working_time = time.strptime(ini_file["업무종료시간"], "%H:%M:%S")

# 데이터 추출 패턴
pattern = r"\d{4}[-]\d{2}[-]\d{2}[ ]\d{2}[:]\d{2}[:]\d{2}[ ]\d+[.]\d+"

# 남길 소수 자릿수 
round_num = int(ini_file["소수자릿수"])

# 엑셀 시작 위치
start_row = int(ini_file["시작행"])
start_col = int(ini_file["시작열"])
column_headname = []
for i in range(7):
    column_headname.append(ini_file["머리말" + str(i + 1)])



def num2excelcol(num):
    
    temp = []
    
    while True:
        num, r = divmod(num, 26)
    
        if r == 0:
            r = 26
            num -= 1

        temp.append(chr(r + 64))

        if num < 26:
            if not num == 0:
                temp.append(chr(num + 64))
            
            col = []
            for i in range(len(temp)-1,-1,-1):
                col.append(temp[i])
            
            return "".join(col)



def average_list(db, num):

    return round(sum(db)/len(db),num)


        
def make_xlsx(data, file_dir):

    wb = openpyxl.Workbook()
    ws = wb.active

    xlsx_file = file_dir[:-3] + "xlsx"

  
    for i in range(5):
        ws.column_dimensions[num2excelcol(start_col + i)].width = 15
        ws.cell(column=start_col + i, row=start_row, value=column_headname[i])
        ws.cell(column=start_col + i, row=start_row).alignment=Alignment(horizontal="center",vertical="center")
        ws.cell(column=start_col + i, row=start_row).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="double"))


    temp1 = []
    temp2 = []
    temp3 = []
    temp4 = []
    for i in range(len(data)):
        for j in range(5):
            ws.cell(column=start_col + j, row=start_row + i + 1, value=data[i][j])
            ws.cell(column=start_col + j, row=start_row + i + 1).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(column=start_col + j, row=start_row + i + 1).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        
            temp1.append(data[i][1])
            temp2.append(data[i][2])
            temp3.append(data[i][3])
            temp4.append(data[i][4])
            
    new_data = [temp1, temp2, temp3, temp4]

        
        
    for i in range(2):
        ws.cell(column=start_col, row=start_row + len(data) + 1 + i, value=column_headname[i + 5])
        ws.cell(column=start_col, row=start_row + len(data) + 1 + i).alignment=Alignment(horizontal="center",vertical="center")
        if i == 0:
            ws.cell(column=start_col, row=start_row + len(data) + 1 + i).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="double"), bottom=Side(style="thin"))
        else:
            ws.cell(column=start_col, row=start_row + len(data) + 1 + i).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        for j in range(4):
            if i == 0:
                ws.cell(column=start_col + 1 + j, row=start_row + len(data) + 1 + i, value=average_list(new_data[j], round_num))
                ws.cell(column=start_col + 1 + j, row=start_row + len(data) + 1 + i).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="double"), bottom=Side(style="thin"))
            else:
                ws.cell(column=start_col + 1 + j, row=start_row + len(data) + 1 + i, value=max(new_data[j]))
                ws.cell(column=start_col + 1 + j, row=start_row + len(data) + 1 + i).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

            ws.cell(column=start_col + 1 + j, row=start_row + len(data) + 1 + i).alignment=Alignment(horizontal="center",vertical="center")
            

                
    wb.save(filename = xlsx_file)

    return 0



def txt_analysis(txt_file):

    file = open(txt_file, "r")

    line_num = len(file.readlines())
    file.seek(0)
    

    # 필요한 데이터만 추출
    ori_db = []
    for i in range(line_num):
        temp = re.findall(pattern,file.readline().strip())

        if not len(temp) == 0:
            ori_db.append(temp[0])
            
    ori_db.sort()


    # 데이터 분리
    new_db = []
    for db in ori_db:
        new_db.append(db.split(" "))


    # 계산
    used_result = []
    for i in range(len(new_db)):

        if i == 0:
            current_date = new_db[i][0]
            temp_work = []
            temp_nonwork = []

        if not current_date == new_db[i][0]:

            work_average = average_list(temp_work, round_num)
            work_max = max(temp_work)
            nonwork_average = average_list(temp_nonwork, round_num)
            nonwork_max = max(temp_nonwork)

            used_result.append([current_date, work_average, work_max, nonwork_average, nonwork_max])

            temp_work = []
            temp_nonwork = []

            current_date = new_db[i][0]
            

        current_time = time.strptime(new_db[i][1], "%H:%M:%S")

        if current_time >= start_working_time and current_time <= end_working_time:
            
            temp_work.append(float(new_db[i][2]))
            
        else:
            
            temp_nonwork.append(float(new_db[i][2]))


        if i == len(new_db) - 1:
            work_average = average_list(temp_work, round_num)
            work_max = max(temp_work)
            nonwork_average = average_list(temp_nonwork, round_num)
            nonwork_max = max(temp_nonwork)

            used_result.append([current_date, work_average, work_max, nonwork_average, nonwork_max])

              
    return used_result



def search_ext(dirname, file_ext, is_full_dir):

    file_list = []
    file_ext = "." + file_ext         
    file_name = os.listdir(dirname)
        
    for file in file_name:
        ext = os.path.splitext(file)[-1]
        
        if ext.upper() == file_ext.upper():
            if is_full_dir == 1:
                if dirname == ".\\":
                    full_dir = os.path.abspath(os.curdir) + "\\" + file
                else:
                    full_dir = os.path.abspath(dirname) + "\\" + file

                file_list.append(full_dir)
                
            else:
                file_list.append(file)
           
    return file_list


        
def main():

    print("작업을 시작합니다.\n")

    # .txt 확장자를 가진 파일 목록 생성
    file_list = []
    full_file_list = []

    file_list += search_ext(base_dir, ext, 0)
    full_file_list += search_ext(base_dir, ext, 1)

    if len(file_list) == 0:
        print(base_dir + "에 txt 파일이 존재하지 않습니다.")
        print("프로그램을 종료합니다.")               
        time.sleep(2)        
        sys.exit()

    print("대상 파일 : " + str(len(file_list)) + "개\n")

    for file in full_file_list:
        
        data = txt_analysis(file)
        
        make_xlsx(data, file)
        
        
    print("모든 작업을 완료했습니다.")
    time.sleep(2)        
    sys.exit()

    
    
if __name__ == '__main__':
    sys.exit(main())
